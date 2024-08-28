import constants
import re
import requests
import time
import pandas
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from urllib3.exceptions import NewConnectionError, MaxRetryError

# Путь к хром-драйверу указывается на каждом локальном компьютере отдельно
chrome_driver_path = "/usr/bin/chromedriver"
driver = webdriver.Chrome(service=Service(chrome_driver_path))


def get_chapter_values(driver=driver):
    """Собираем все значения разделов выпадающего списка."""
    html_content = fetch_html_content(constants.url)
    if not html_content:
        return
    soup = BeautifulSoup(html_content, "html.parser")
    dropdown_list = soup.find("select", id=constants.dropdown_list_id)
    values_dict = {}
    for option in dropdown_list.find_all("option"):
        value = option["value"]
        text = option.text.strip()
        values_dict[value] = text
    if constants.max_value in values_dict:
        del values_dict[constants.max_value]
    view_chapters(values_dict, driver)


def fetch_html_content(url):
    """Проверяем ошибки при выполнении запроса."""
    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.content
    except (requests.exceptions.HTTPError, requests.exceptions.ConnectionError,
            requests.exceptions.Timeout, NewConnectionError, MaxRetryError,
            requests.exceptions.RequestException) as error:
        print(f"Ошибка запроса {error}")
        return None


def view_chapters(values_dict, driver):
    """Выполняем просмотр страниц заданий."""
    driver.maximize_window()
    # Заходим на главную страницу
    driver.get(constants.url)
    # Идем циклом по всем разделам:
    checkboxes_dict = {}
    for value in values_dict.items():
        # Выбираем раздел выпадающего списка
        print(f"Обработка раздела {value[1]}")
        get_chapter(value[0])
        # Получаем название каждого чекбокса
        checkboxes = get_table().find_elements(By.TAG_NAME, "input")
        checkboxes_names = []
        get_check_box_names(checkboxes, checkboxes_names)
        # Добавляем в словарь Id раздела и список с чебоксами для каждого раздела
        checkboxes_dict[value[0], value[1]] = checkboxes_names
    print("Собраны данные для парсинга")
    # Запускаем парсер по собранным данным
    parsing_chapters(checkboxes_dict)
    driver.quit()


def get_check_box_names(checkboxes, checkboxes_names):
    """Получение имен чекбоксов."""
    for checkbox in checkboxes:
        if checkbox.get_attribute("type") == "checkbox":
            checkboxes_names.append(checkbox.get_attribute("Name"))
        time.sleep(constants.sleep_time)


def get_table():
    """Получаем таблицу с чекбоксами."""
    table = driver.find_element(By.ID, constants.table_id)
    time.sleep(constants.sleep_time)
    return table


def get_chapter(value):
    """Выбираем раздел выпадающего списка."""
    dropdown_list = driver.find_element(By.ID, constants.dropdown_list_id)
    dropdown_list.click()
    option = dropdown_list.find_element(By.XPATH,
                                        f"//option[@value='{value}']")
    option.click()
    time.sleep(constants.sleep_time)


def parsing_chapters(checkboxes_dict):
    """Парсинг данных со страницы подраздела с задачами."""
    print("Сбор данных")
    for id in checkboxes_dict:
        # Формируем ссылку для каждого отдельного подраздела
        for name in checkboxes_dict[id]:
            url = f"https://kpolyakov.spb.ru/school/ege/gen.php?action=viewAllEgeNo&egeId={id[0]}&{name}=on"
            driver.get(url)
            soup = BeautifulSoup(driver.page_source, "lxml")
            subject = id[1]
            # Находим подтему
            subtopic = parse_subtopic(soup)
            table = soup.find_all("td", {"class": "topicview"})
            # Находим задачи
            for td in table:
                # Убираем лишний элемент в коде
                clear_script(td)
                # Получаем ссылку на изображения
                image = get_image(td)
                # Получаем номер задачи
                task_number = get_task_number(td)
                # Получаем текст задачи
                task_text = get_task_text(td)
                # Получаем ответ задачи
                task_asnwer = get_task_answer(table, soup, task_number)
                result = {
                    "Номер задания": task_number,
                    "Автор": "ЕГЭ-2024",
                    "Условие": f"{task_text} \n{image}",
                    "Ответ": task_asnwer,
                    "Тема": subject,
                    "Подтема": subtopic
                }
                # Отправляем данные в Excel
                push_data_to_excel_file(result)
            print(f"Обработан раздел {subtopic}")
        print(f"Обработана тема {subject}")
    print("Все данные собраны")


def parse_subtopic(soup):
    """Находим подтему."""
    subtopics = soup.find_all("p")
    for subtopic in subtopics:
        if "Раздел" in subtopic.text:
            subtopic = subtopic.find("b").text
            return subtopic


def clear_script(td):
    """Убираем элемент script из HTML-кода."""
    script = td.find("script")
    script.decompose()


def get_image(td):
    """Получаем изображение."""
    image_url = " "
    img_tag = td.find("img")
    if td.find("img"):
        image_src = img_tag["src"]
        img_tag.decompose()
        url_prefix = "https://kpolyakov.spb.ru/"
        new_path = image_src.replace("../../", "")
        image_url = url_prefix + new_path
        return image_url
    return image_url


def get_task_number(td):
    """Получаем номер задания."""
    text_task = td.get_text().strip()
    pattern = r"(?:#)?(\d+)"
    match = re.search(pattern, text_task)
    if match:
        number = match.group(1)
        return number


def get_task_text(td):
    """Получаем текст задачи."""
    text_task = td.get_text().strip()
    pattern = r'(?:#)?\(([^)]+)\)'
    text_task_result = re.sub(pattern, "", text_task).strip()
    return text_task_result


def get_task_answer(table, soup, task_number):
    """Находим ответ к задаче."""
    div_tags = soup.find_all("div", {"class": "hidedata"})
    for div_tag in div_tags:
        if div_tag["id"] == task_number:
            return div_tag.get_text()


def push_data_to_excel_file(result):
    """Отправляем данные в excel_file."""
    file_path = "Информатика.xlsx"
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        df_new = pandas.DataFrame([result])
        for _, row in df_new.iterrows():
            ws.append(row.tolist())
        max_width = 60
        for column in ws.columns:
            length = max(len(str(cell.value)) for cell in column)
            adjusted_width = min(length + 2, max_width)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        for cell in ws.iter_rows(min_row=2, values_only=False):  # Игнорируем заголовок
            for c in cell:
                if isinstance(c.value, str) and len(
                        str(c.value)) > 50:
                    row_index = c.row
                    column_index = c.column
                    # Создаем объект RowDimension
                    row_dim = ws.row_dimensions[row_index]
                    # Устанавливаем высоту строки в 7 см
                    row_dim.height = 7 * 20  # 1 см = 20 единицам в Excel
                    # Устанавливаем ширину столбца
                    ws.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = 30
                    ws.cell(row=row_index, column=column_index).alignment = Alignment(
                        wrap_text=True,
                        vertical='center',
                        shrink_to_fit=False
                    )

        wb.save(file_path)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        df_all = pandas.DataFrame([result])
        ws.append(list(df_all.columns))
        for _, row in df_all.iterrows():
            ws.append(row.tolist())
        # Форматирование ячеек с длинным текстом
        for cell in ws.iter_rows(min_row=2, values_only=False):
            for c in cell:
                if isinstance(c.value, str) and len(str(c.value)) > 50:
                    row_index = c.row
                    column_index = c.column
                    # Создаем объект RowDimension
                    row_dim = ws.row_dimensions[row_index]
                    # Устанавливаем высоту строки в 7 см
                    row_dim.height = 7 * 20  # 1 см = 20 единицам в Excel
                    # Устанавливаем ширину столбца
                    ws.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = 30
                    ws.cell(row=row_index, column=column_index).alignment = Alignment(
                        wrap_text=True,
                        vertical='center',
                        shrink_to_fit=False
                    )

        wb.save(file_path)


def main():
    """Основная функция."""
    get_chapter_values()


if __name__ == "__main__":
    main()
